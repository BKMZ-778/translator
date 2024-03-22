import openpyxl
import pandas as pd
from openpyxl import Workbook
import operator
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


df = pd.read_excel('test_words.xlsx', sheet_name=0, engine='openpyxl', usecols='A')
print(df)
df_catalog = pd.read_excel('КАТАЛОГ для кодов.xlsx', sheet_name='Проверка кодов', engine='openpyxl', usecols='A, B', dtype=str)
print(df_catalog)
df['santanses'] = df['santanses'].astype(str).str.lower()
df_catalog['word'] = df_catalog['word'].astype(str).str.lower()
df_catalog['code'] = df_catalog['code'].astype(str)
sant_list = df['santanses'].drop_duplicates().to_list()
catal_list_dict = df_catalog.drop_duplicates().to_dict('records')
catal_list_dict = sorted(catal_list_dict, reverse=True, key=lambda s: len(s['word']))
dict_result = {}
for row in sant_list:
    in_string_list = []
    for row_cat in catal_list_dict:
        if row_cat['word'] in row:
            in_string_list.append(row_cat)
    a = str(in_string_list).replace("{'word': ", '')
    a = a.replace("'code': ", '')
    a = a.replace("'", '')
    a = a.replace("}", '')
    a = a.replace("]", '')
    a = a.replace("[", '')
    dict_result[row] = a
    in_string_list = []
append_list = []
for el in dict_result:
    if dict_result[el] == []:
        append_list.append(el)
df_dict_result = pd.DataFrame(dict_result, index=[0]).reset_index().transpose()
df_dict_result[['word', 'ect']] = df_dict_result[0].str.split(',', 1, expand=True)
df_dict_result = df_dict_result.drop(0, axis=1)
df_dict_result[['1', '2']] = df_dict_result['ect'].str.split(',', 1, expand=True)
df_dict_result = df_dict_result.drop("ect", axis=1)
df_dict_result[['word2', '4']] = df_dict_result['2'].str.split(',', 1, expand=True)
df_dict_result = df_dict_result.drop("2", axis=1)
df_dict_result[['2', 'ect']] = df_dict_result['4'].str.split(',', 1, expand=True)
df_dict_result = df_dict_result.drop("4", axis=1)
writer = pd.ExcelWriter('df_dict_result.xlsx', engine='xlsxwriter')
df_dict_result.to_excel(writer, sheet_name='Sheet1')
for column in df_dict_result:
    column_width = max(df_dict_result[column].astype(str).map(len).max(), len(column))
    col_idx = df_dict_result.columns.get_loc(column)
    writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
    writer.sheets['Sheet1'].set_column(0, 3, 10)
    writer.sheets['Sheet1'].set_column(1, 3, 20)
    writer.sheets['Sheet1'].set_column(2, 3, 20)
    writer.sheets['Sheet1'].set_column(3, 3, 20)
    writer.sheets['Sheet1'].set_column(4, 3, 30)
    writer.sheets['Sheet1'].set_column(5, 3, 20)
writer.save()
df_to_append = pd.DataFrame(append_list, columns=['points'])
df_to_append = df_to_append.sort_values(by='points')

wb = openpyxl.load_workbook('df_dict_result.xlsx')
ws = wb.active

ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToHeight = False
cm = int(1 / 4)
ws.page_margins = PageMargins(left=cm, right=cm, top=cm, bottom=cm)
i = 0
al = Alignment(horizontal="left", vertical="top")
for cell in ws['A']:
    i += 1
    cell.alignment = al
    print(cell.value)

writer = pd.ExcelWriter('points.xlsx', engine='xlsxwriter')
df_to_append.to_excel(writer, sheet_name='Sheet1', index=False)
writer.save()

