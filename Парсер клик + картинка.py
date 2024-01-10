import webbrowser
import pandas as pd
from tkinter import filedialog
import tkinter
from tkinter import *
import time
from threading import Thread
import keyboard
import pyautogui as pag
import pyperclip
import tkinter.messagebox as mb
import openpyxl
import mss.tools
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment

"""sr = speech_recognition.Recognizer()
sr.pause_threshold = 1"""

#fileName = filedialog.askopenfilename()
fileName = filedialog.askopenfilename()

df_list = pd.read_excel(fileName, sheet_name=0, engine='openpyxl', header=None, usecols='L, N', skiprows=1)

df_base_to_work = pd.read_excel('C:/Users/User/Desktop/РЕЕСТРЫ/БАЗА подтяжек.xlsx', sheet_name=0, engine='openpyxl', header=None, usecols='A, C')
df_merged = pd.merge(df_list, df_base_to_work, how='left', left_on=11, right_on=0)
df_to_append = df_merged[df_merged[2].isnull()].drop_duplicates(11)
df_to_append.columns = range(df_to_append.shape[1])
print(df_to_append)
print(len(df_to_append))
china_names = df_to_append[df_to_append.columns[0]].tolist()
urls = df_to_append[df_to_append.columns[1]].tolist()
print(china_names)
print(urls)

label = 1
label_3 = 1
def shift_click():
    global running
    if running == True:
        running = False
    else:
        running = True


def window():
    global label
    top = tkinter.Tk()
    Button = tkinter.Button(top, text="Start", width=8, command=shift_click, height=2)
    Button.configure(font=('Times', 20))
    lable_main = Label(top, text='Парсер из заголовков ссылки (Яндекс.браузер)', fg='black')
    lable_main.pack()
    Button.pack()
    label_0 = Label(top, text='№ Ссылки:', fg='black')
    label_0.pack()
    label = Label(top, text='', fg='black', font=('Times', 30))
    label.pack()
    label_2 = Label(top, text='время задержки:', fg='black')
    label_2.pack()
    entry = Entry(top)
    entry.pack()
    def get_profit():
        global profit
        profit = int(entry.get())
    button_take_time = tkinter.Button(top, text="Установить", command=get_profit)
    button_take_time.configure(font=('Times', 10))
    button_take_time.pack()
    top.mainloop()

keyboard.add_hotkey('shift', shift_click)
running = False

t = Thread(target=window)
t.start()


i = 0
list_itemDescriptions = []
wb = Workbook()
ws = wb.active
for url, name in zip(urls, china_names):
    if running == False:
        while running == False:
            # ожидаем повторного нажатия кнопочки
            time.sleep(2)
    if running == True:
        # выполняем основную полезную работу программы
        i += 1
        print(i)
        label.config(text=i, fg='black', font=("Times", 20))
        try:
            webbrowser.open(url)
            time.sleep(profit)
            pag.click(x=580, y=55, button='right')
            time.sleep(1)
            pag.click(x=605, y=90)
            if 'Кулинарный инструмент' in pyperclip.paste():
                pag.click(x=155, y=419)
                time.sleep(profit)
                pag.click(x=536, y=58, button='right')
                time.sleep(1)
                pag.click(x=570, y=92)
            pag.screenshot(f'{i}.png', region=(200, 200, 900, 800))
            """with mss.mss() as sct:
                # The screen part to capture
                monitor = {"top": 400, "left": 400, "width": 400, "height": 400}
                output = f"{i}.png"
                # Grab the data
                sct_img = sct.grab(monitor)
                # Save to the picture file
                mss.tools.to_png(sct_img.rgb, sct_img.size, output=output)
                print(output)"""
            time.sleep(1)
            list_itemDescriptions.append(pyperclip.paste())
            img = openpyxl.drawing.image.Image(f'{i}.png')
            img.height = 190 # insert image height in pixels as float or int (e.g. 305.5)
            img.width = 430 # insert image width in pixels as float or int (e.g. 405.8)
            row_num = i
            cell_addr = f"A{row_num}"
            img.anchor = cell_addr
            ws.add_image(img)
            ws[f"C{row_num}"] = name
            ws[f"C{row_num}"].alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
            ws[f"D{row_num}"].hyperlink = url
            ws[f"D{row_num}"].value = url
            ws[f"D{row_num}"].style = "Hyperlink"
            ws[f"D{row_num}"].alignment = Alignment(horizontal='center')
            ws[f"B{row_num}"] = pyperclip.paste()
            ws[f"B{row_num}"].alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
            ws.row_dimensions[row_num].height = int(180 * .8)
            ws.column_dimensions["A"].width = int(300 * .2)
            pag.click(x=189, y=15)
            pag.hotkey('ctrl', 'w')
        except EXCEPTION:
            list_itemDescriptions.append('Error')

wb.save(f'{fileName}-translate_table.xlsx')
msg = "Смотри файл 'output' в папке с проектом"
mb.showinfo("важная информация", msg)
print('Готово')
input()

