import pandas as pd
import tkinter as tk
from tkinter import *
from tkinter import filedialog
import tkinter.messagebox as mb
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import sqlite3 as sl
from openpyxl.utils.dataframe import dataframe_to_rows
import random
import numpy as np


def start_ozon():
    fileName = filedialog.askopenfilename()
    df_ozon_file = pd.read_excel(fileName, sheet_name=0, engine='openpyxl', header=None, usecols='L,M,N', skiprows=1)
    df_ozon_file = df_ozon_file.rename(columns={11: 'bad_description', 12: 'price', 13: 'link'})
    print(df_ozon_file)
    df_baza_ozon = pd.read_excel('C:/Users/User/Desktop/РЕЕСТРЫ/БАЗА ОЗОН ОПИСАНИЯ.xlsx', sheet_name=0, engine='openpyxl', header=None, usecols='A,B', skiprows=1)
    df_baza_ozon = df_baza_ozon.rename(columns={0: 'bad_description', 1: 'good_description'})
    print(df_baza_ozon)
    df_merged = pd.merge(df_ozon_file, df_baza_ozon, how='left', left_on='bad_description', right_on='bad_description')

    df_to_translate = df_merged.loc[df_merged['good_description'].isnull()]
    df_to_translate['good_description'] = df_to_translate['bad_description']
    df_to_translate = df_to_translate[['bad_description', 'good_description', 'price', 'link']]
    df_to_translate = df_to_translate.drop_duplicates(subset='good_description', keep='first').sort_values(by='good_description')
    print(df_to_translate)
    """writer = pd.ExcelWriter(f'{fileName}-На перевод.xlsx', engine='openpyxl')
    df_to_translate.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()

    wb = openpyxl.load_workbook(f'{fileName}-На перевод.xlsx')
    ws = wb.active
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 75
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 9
    wb.save(f'{fileName}-На перевод.xlsx')"""

    wb = openpyxl.load_workbook(f'{fileName}')
    ws = wb.active
    ws.title = "РЕЕСТР"
    wb.save(f'{fileName}')
    wb2 = openpyxl.load_workbook(f'{fileName}')
    ws2 = wb2.create_sheet("Sheet_A")
    ws2.title = "На_перевод"

    rows = dataframe_to_rows(df_to_translate, index=False)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=value)
    ws2.column_dimensions['A'].width = 9
    ws2.column_dimensions['B'].width = 75
    ws2.column_dimensions['C'].width = 9
    ws2.column_dimensions['D'].width = 9
    wb2.save(f'{fileName}-На перевод.xlsx')


    msg = "Готово!"
    mb.showinfo("Информация", msg)

def add_tobaza():
    fileName = filedialog.askopenfilename()
    df_to_append = pd.read_excel(f'{fileName}', sheet_name='На_перевод')
    df_base = pd.read_excel('C:/Users/User/Desktop/РЕЕСТРЫ/БАЗА ОЗОН ОПИСАНИЯ.xlsx')
    df_base_updated = pd.concat([df_base, df_to_append], axis=0).drop_duplicates(subset='bad_description', keep='last')
    writer = pd.ExcelWriter('C:/Users/User/Desktop/РЕЕСТРЫ/БАЗА ОЗОН ОПИСАНИЯ.xlsx', engine='openpyxl')
    df_base_updated.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()

    df_ozon_file = pd.read_excel(fileName, sheet_name=0, engine='openpyxl', header=None, usecols='L', skiprows=1)
    print(df_ozon_file)
    df_ozon_file = df_ozon_file.rename(columns={11: 'bad_description'})
    df_merged = pd.merge(df_ozon_file, df_base_updated, how='left', left_on='bad_description', right_on='bad_description')
    print(df_merged)
    wb = openpyxl.load_workbook(fileName)
    ws = wb['РЕЕСТР']
    ws.insert_cols(13)
    i = 2
    for row in df_merged['good_description']:
        ws[f"M{i}"].value = row
        i += 1
    ws['L1'].value = 'СТАРОЕ НАИМЕНОВАНИЕ'
    ws['M1'].value = 'наименованиетовара/名称/俄文/中文'

    if 'На_перевод' in wb.sheetnames:
        wb.remove(wb['На_перевод'])
    if 'с_картинкой' in wb.sheetnames:
        wb.remove(wb['с_картинкой'])
    wb.save(f'{fileName}-ГОТОВ.xlsx')

    msg = "Обновленно!"
    mb.showinfo("Информация", msg)

def start_LD():
    fileName = filedialog.askopenfilename()
    df_LD_file = pd.read_excel(fileName, sheet_name=0, engine='openpyxl', header=None, usecols='L,M,N,Y', skiprows=1)
    df_LD_file = df_LD_file.rename(columns={11: 'china_description', 12: 'price', 13: 'link', 24: 'SKU'})
    df_baza_LD = pd.read_excel('C:/Users/User/Desktop/РЕЕСТРЫ/БАЗА ЛД.xlsx', sheet_name=0, engine='openpyxl', header=None, usecols='A,C', skiprows=1)
    df_baza_LD = df_baza_LD.rename(columns={0: 'SKU', 2: 'good_description'})
    df_merged = pd.merge(df_LD_file, df_baza_LD, how='left', left_on='SKU', right_on='SKU')
    print(df_merged)
    df_to_translate = df_merged.loc[df_merged['good_description'].isnull()]
    print(df_to_translate)
    df_to_translate = df_to_translate[['china_description', 'good_description', 'price', 'link', 'SKU']]
    df_to_translate = df_to_translate.drop_duplicates(subset='SKU', keep='first').sort_values(by='china_description')
    print(df_to_translate)

    """writer = pd.ExcelWriter(f'{fileName}-На перевод.xlsx', engine='openpyxl')
    df_to_translate.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()

    wb = openpyxl.load_workbook(f'{fileName}-На перевод.xlsx')
    ws = wb.active

    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 9

    """

    wb = openpyxl.load_workbook(fileName)
    ws = wb.active
    ws.title = "РЕЕСТР"
    wb.save(fileName)
    wb2 = openpyxl.load_workbook(fileName)
    ws2 = wb2.create_sheet("Sheet_A")
    ws2.title = "На_перевод"

    rows = dataframe_to_rows(df_to_translate, index=False)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=value)
    ws2.column_dimensions['A'].width = 9
    ws2.column_dimensions['B'].width = 9
    ws2.column_dimensions['C'].width = 9
    ws2.column_dimensions['D'].width = 9
    len_sheet = ws2.max_row
    ws2.move_range(f"D1:D{len_sheet}", cols=10)
    ws2.move_range(f"E1:E{len_sheet}", cols=7)
    wb2.save(f'{fileName}-На перевод.xlsx')

    msg = "Готово!"
    mb.showinfo("Информация", msg)

def LD_work():
    fileName = filedialog.askopenfilename()
    df_baza_LD = pd.read_excel('C:/Users/User/Desktop/РЕЕСТРЫ/БАЗА ЛД.xlsx', sheet_name=0, engine='openpyxl', usecols='A,B,C,D')
    wb = openpyxl.load_workbook(fileName)
    ws2 = wb['На_перевод']
    df_translate_table = pd.read_excel(fileName, sheet_name='с_картинкой', engine='openpyxl', header=None, usecols='B')
    print(df_translate_table)
    len_sheet = ws2.max_row
    ws2.move_range(f"C1:C{len_sheet}", cols=1)
    ws2.move_range(f"B1:B{len_sheet}", cols=1)
    ws2.move_range(f"A1:A{len_sheet}", cols=1)

    ws2.move_range(f"L1:L{len_sheet}", cols=-11)
    ws2.move_range(f"N1:N{len_sheet}", cols=-9)


    i = 2
    for row in df_translate_table[1]:
        ws2[f"C{i}"].value = row
        i += 1
    wb.save(fileName)
    df_to_append = pd.read_excel(fileName, sheet_name='На_перевод')
    df_baza_LD_updated = pd.concat([df_baza_LD, df_to_append], axis=0).drop_duplicates(subset='SKU', keep='last')
    writer = pd.ExcelWriter('C:/Users/User/Desktop/РЕЕСТРЫ/БАЗА ЛД.xlsx', engine='openpyxl')
    df_baza_LD_updated.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()

    df_LD_registry = pd.read_excel(fileName, sheet_name='РЕЕСТР', engine='openpyxl', header=None, usecols='Y', skiprows=1)
    df_LD_registry = df_LD_registry.rename(columns={24: 'SKU'})
    df_merged = pd.merge(df_LD_registry, df_baza_LD_updated, how='left', left_on='SKU',
                         right_on='SKU')
    print(df_merged)
    wb = openpyxl.load_workbook(fileName)
    ws = wb['РЕЕСТР']
    ws.insert_cols(13)
    i = 2
    for row in df_merged['good_description']:
        ws[f"M{i}"].value = row
        i += 1

    if 'На_перевод' in wb.sheetnames:
        wb.remove(wb['На_перевод'])
    if 'с_картинкой' in wb.sheetnames:
        wb.remove(wb['с_картинкой'])
    wb.save(f'{fileName}-ГОТОВ.xlsx')

    msg = "Готово!"
    mb.showinfo("Информация", msg)

def passport_check():
    fileName = filedialog.askopenfilename()

    df = pd.read_excel(fileName, usecols='A, O, P', engine='openpyxl', header=None, skiprows=1)
    print(df)
    df['трек'] = df[0]
    df['Серия паспорта'] = df[14]
    df['Номер паспорта'] = df[15]

    df['Серия паспорта'] = pd.to_numeric(df['Серия паспорта'], errors='coerce').fillna(0).astype('int64')
    df['Номер паспорта'] = pd.to_numeric(df['Номер паспорта'], errors='coerce').fillna(0).astype('int64')
    print(type(df['Серия паспорта'].values[1]))
    print(df['Серия паспорта'])

    df['Серия паспорта'] = df['Серия паспорта'].astype('str')
    df['Номер паспорта'] = df['Номер паспорта'].astype('str')
    df_to_baza = df.drop_duplicates()
    df_to_baza = df_to_baza[['Серия паспорта', 'Номер паспорта']]
    df_to_baza = df_to_baza.rename(columns={'Серия паспорта': 'series', 'Номер паспорта': 'number'})

    con = sl.connect(r'C:\Users\User\PycharmProjects\Переводчик\PASS_BAZA.db')
    with con:
        baza = con.execute("select count(*) from sqlite_master where type='table' and name='pass'")
        for row in baza:
            # если таких таблиц нет
            if row[0] == 0:
                # создаём таблицу
                with con:
                    con.execute("""
                                CREATE TABLE pass (
                                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                series VARCHAR(4),
                                number VARCHAR(6)
                                );
                            """)
                    con.execute("""CREATE INDEX index_pass ON pass (series, number)""")
        df_to_baza.to_sql('pass', con, if_exists='replace', index=False)

        names = [description[0] for description in con.execute("Select * from pass").description]
    print(names)
    con.commit()

    # drop table
    # con = sqlite3.connect(r'C:\Users\Илья\PycharmProjects\\Переводчик\save_pandas.db')
    # with con:
    #    con.executescript('drop table if exists pass')

    with con:
        # Query for INNER JOIN
        sql = '''SELECT PASS_BAZA.PASSP_SERIES, PASS_BAZA.PASSP_NUMBER 
        FROM PASS_BAZA 
        INNER JOIN pass
        ON PASS_BAZA.PASSP_SERIES = pass.series
        AND PASS_BAZA.PASSP_NUMBER = pass.number'''
        df_finish = pd.read_sql(sql, con)

    con.commit()
    con.close()

    df['ID'] = df['Серия паспорта'] + df['Номер паспорта']
    df_finish['ID'] = df_finish['PASSP_SERIES'] + df_finish['PASSP_NUMBER']
    df_merged = pd.merge(df, df_finish, how='left', left_on='ID', right_on='ID')
    df_merged_to_select = df_merged[df_merged['PASSP_SERIES'].notnull()]
    df_merged_to_select = df_merged_to_select.drop_duplicates()
    df_merged_to_select = df_merged_to_select.dropna(how='any', axis=0)
    df_merged_to_select = df_merged_to_select[df_merged_to_select['PASSP_SERIES'] != '0']
    redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    print(df_merged_to_select)
    if not df_merged_to_select.empty:
        wb = openpyxl.load_workbook(fileName)
        ws = wb.active
        parcel_list = df_merged_to_select[0].to_list()
        i = 0
        for cell in ws['A']:
            i += 1
            if cell.value in parcel_list:
                ws[f'O{i}'].fill = redFill
                ws[f'P{i}'].fill = redFill
                print('ok')
        wb.save(fileName)

        #writer = pd.ExcelWriter(f'{fileName} - список недействительных.xlsx', engine='openpyxl')
        #df_merged_to_select.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
        #writer.save()

    wb = openpyxl.load_workbook(fileName)
    try:
        ws = wb['РЕЕСТР']
    except:
        ws = wb.active
    # df_to_check = pd.read_excel(fileName, usecols='L', engine='openpyxl', header=None, skiprows=1)
    df_exclude = pd.read_excel('exclude.xlsx', engine='openpyxl', converters={'Исключить': str, 'Убрать': str,
                                                                              'Заменить': str, 'ЗаменитьНа': str})

    exclude_list = df_exclude['Исключить'].to_list()
    change_list = df_exclude['Убрать'].to_list()
    change_list = [item for item in change_list if not (pd.isnull(item)) is True]
    change_list_2 = df_exclude['Заменить'].to_list()
    change_list_2 = [item for item in change_list_2 if not (pd.isnull(item)) is True]
    change_list_3 = df_exclude['ЗаменитьНа'].to_list()
    change_list_3 = [item for item in change_list_3 if not (pd.isnull(item)) is True]

    i = 0
    for cell in ws['L']:
        i += 1
        numb = 0
        for item_ch in change_list:
            try:
                if item_ch in cell.value.lower():
                    print(cell.value)
                    cell.value = cell.value.lower().replace(item_ch, '')
                    print(cell.value)
            except:
                pass
        for item_ch_2 in change_list_2:
            try:
                if item_ch_2.lower() in cell.value.lower():
                    cell.value = cell.value.lower().replace(item_ch_2, change_list_3[numb])
                    print(item_ch_2)
            except:
                pass
    for cell in ws['L']:
        for item in exclude_list:
            try:
                if item in cell.value.lower():
                    cell.fill = redFill
                    print(item)
            except:
                pass

            numb += 1
    i = 0
    for cell in ws['O']:
        print(cell.value)
        i += 1
        if cell.value is None:
            cell.fill = redFill

    wb.save(fileName)

    msg = "ГОТОВО!"
    mb.showinfo("ИНФО", msg)


def pudo():
    fileName = filedialog.askopenfilename()
    df_start = pd.read_excel(fileName, sheet_name=0, engine='openpyxl', header=None, usecols='W, AT', skiprows=1)
    df_start = df_start.rename(columns={22: 'weight', 45: 'chin'})
    df_tosql = df_start.drop(['weight'], axis=1).drop_duplicates()
    con = sl.connect('TRANSLATE.db')
    with con:
        df_tosql.to_sql('temp_table', con, if_exists='replace', index=False)
        query_indx = """CREATE INDEX index_chin on temp_table (chin)"""
        print('index created')
        con.execute(query_indx)
        query_join = """SELECT temp_table.chin, transl_cainiao.rus
                        FROM temp_table
                        LEFT JOIN transl_cainiao
                        ON temp_table.chin = transl_cainiao.chin
                        group by temp_table.chin"""

        df_joined = pd.read_sql(query_join, con)
        print(df_joined)
        df_merged = pd.merge(df_start, df_joined, how='left', left_on='chin', right_on='chin')

        list_values = ["Аксессуар из пластика", "Украшение", "Игрушка"]
        len_df = len(df_merged)
        df = df_merged.loc[df_merged['rus'].isna()]['rus'].apply(lambda x: random.choices(list_values, k=len_df)[0])
        df_merged.update(df)
        df_weight_big = df_merged.loc[df_merged['weight'] >= 3]
        df_weight_big = df_weight_big.loc[(df_merged['rus'] == 'Аксессуар из пластика') |
                                       (df_merged['rus'] == 'Украшение') | (df_merged['rus'] == 'Игрушка')]
        df_weight_big['rus'] = ''
        print(df_weight_big)
        writer = pd.ExcelWriter(f'df_weight_big.xlsx', engine='openpyxl')
        df_merged.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()

        df_merged.update(df_weight_big)

        writer = pd.ExcelWriter(f'df_merged.xlsx', engine='openpyxl')
        df_merged.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()

        wb = openpyxl.load_workbook(fileName)
        ws = wb.active
        i = 2
        for row in df_merged['rus']:
            ws[f"AM{i}"].value = row
            i += 1
        wb.save(f'{fileName}-ГОТОВ.xlsx')

        msg = "ГОТОВО!"
        mb.showinfo("ИНФО", msg)


def export_warrings():
    filename = filedialog.askopenfilename()
    df = pd.read_excel(filename, sheet_name=0, engine='openpyxl', header=None, usecols='A, B', skiprows=1, converters={1: str})
    df.columns = ['description', 'tnvedCode']
    print(df)
    df['tnvedCode9'] = df['tnvedCode'].str[:9]
    df['tnvedCode6'] = df['tnvedCode'].str[:6]
    df['tnvedCode4'] = df['tnvedCode'].str[:4]

    df_warnings = pd.read_excel('TNVEDзапрет.xlsx', sheet_name=0, engine='openpyxl', usecols='A, B, C, D', converters={'triger_tnved': str,
                                                                                                                       'exclud': str})

    df_merged_10 = pd.merge(df, df_warnings, how='left', left_on='tnvedCode', right_on='triger_tnved')
    df_merged_9 = pd.merge(df_merged_10, df_warnings, how='left', left_on='tnvedCode9', right_on='triger_tnved')
    df_merged_6 = pd.merge(df_merged_9, df_warnings, how='left', left_on='tnvedCode6', right_on='triger_tnved')
    df_merged_4 = pd.merge(df_merged_6, df_warnings, how='left', left_on='tnvedCode4', right_on='triger_tnved')

    df_merged_4.columns = ['description', 'tnvedCode', 'tnvedCode9', 'tnvedCode6', 'tnvedCode4', 'Триггер_код', 'Описание группы/запрет'
                           , 'Постановление', 'Искл.', 'По 9 знакам', 'Описание группы/запрет_9', 'Постановление_9', 'Искл._9',
                           'По 6 знакам', 'Описание группы/запрет_6', 'Постановление_6', 'Искл._6'
                           , 'По 4 знакам', 'Описание группы/запрет_4', 'Постановление_4', 'Искл._4']
    df = df_merged_4
    df = df.drop('Искл.', axis=1)
    df = df.drop('Искл._6', axis=1)
    df = df.drop('Искл._9', axis=1)
    #df['Искл._4'] = df['Искл._4']

    df['Постановление'] = df['Постановление'].fillna(df['Постановление_9'])
    df['Постановление'] = df['Постановление'].fillna(df['Постановление_6'])
    df['Постановление'] = df['Постановление'].fillna(df['Постановление_4'])

    df['Триггер_код'] = df['Триггер_код'].fillna(df['По 9 знакам'])
    df['Триггер_код'] = df['Триггер_код'].fillna(df['По 6 знакам'])
    df['Триггер_код'] = df['Триггер_код'].fillna(df['По 4 знакам'])

    df['Описание группы/запрет'] = df['Описание группы/запрет'].fillna(df['Описание группы/запрет_9'])
    df['Описание группы/запрет'] = df['Описание группы/запрет'].fillna(df['Описание группы/запрет_6'])
    df['Описание группы/запрет'] = df['Описание группы/запрет'].fillna(df['Описание группы/запрет_4'])

    df = df.drop('tnvedCode9', axis=1)

    writer = pd.ExcelWriter(f'TNVED_запрет_вывоза_результат.xlsx', engine='openpyxl')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()

    df_trigger_words_export = pd.read_excel("trigger_words_export.xlsx", header=None)
    list_of_trigger_words = df_trigger_words_export[0].to_list()
    wb = openpyxl.load_workbook('TNVED_запрет_вывоза_результат.xlsx')
    ws = wb.active
    orangeFill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    pinkFill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')

    i = 0
    for row in ws['B']:
        i += 1
        tnved = str(row.value)
        for triger_word in list_of_trigger_words:
            if triger_word in ws[f'A{i}'].value.lower():
                ws[f'A{i}'].fill = orangeFill
        if ws[f'E{i}'].value is not None and i != 1 and ws[f'E{i}'].value != '3926909709' and ws[f'E{i}'].value != '621133'\
                and ws[f'E{i}'].value != '6914900000' and ws[f'E{i}'].value != '7326909807':
            print(ws[f'E{i}'].value)
            ws[f'A{i}'].fill = redFill
            ws[f'B{i}'].fill = redFill
            ws[f'E{i}'].fill = redFill
        try:
            excludings = ws[f'Q{i}'].value.split(",")
            for exclude in excludings:
                print(exclude)
                if exclude in tnved:
                    ws[f'A{i}'].fill = pinkFill
                    ws[f'B{i}'].fill = pinkFill
                    ws[f'Q{i}'].fill = pinkFill
        except:
            excludings = ws[f'Q{i}'].value
        #print(tnved, excludings)

    wb.save('TNVED_запрет_вывоза_результат.xlsx')

    msg = "ГОТОВО!"
    mb.showinfo("ИНФО", msg)


window = tk.Tk()
window.title('OZON')
window.geometry("400x310+500+300")

button = tk.Button(text="На перевод OZON", width=24, height=2, bg="lightgrey", fg="black", command=start_ozon)
button.configure(font=('hank', 10))

button2 = tk.Button(text="OZON Добавить в базу + обновить реестр", width=35, height=2, bg="lightgrey", fg="black", command=add_tobaza)
button2.configure(font=('hank', 10))

button3 = tk.Button(text="На перевод LD", width=24, height=2, bg="lightgrey", fg="black", command=start_LD)
button3.configure(font=('hank', 10))

button4 = tk.Button(text="LD Добавить в базу + обновить реестр", width=35, height=2, bg="lightgrey", fg="black", command=LD_work)
button4.configure(font=('hank', 10))

button5 = tk.Button(text="Проверка запрещенка + паспорт", width=35, height=2, bg="lightgrey", fg="black", command=passport_check)
button5.configure(font=('hank', 10))

button7 = tk.Button(text="PUDO", width=35, height=2, bg="lightgrey", fg="black", command=pudo)
button7.configure(font=('hank', 10))
button8 = tk.Button(text="ЗАПРЕТ НА ВЫВОЗ", width=35, height=2, bg="lightgrey", fg="black", command=export_warrings)
button8.configure(font=('hank', 10))

button.pack()
button2.pack()
button3.pack()
button4.pack()
button7.pack()
button5.pack()
button8.pack()


window.mainloop()