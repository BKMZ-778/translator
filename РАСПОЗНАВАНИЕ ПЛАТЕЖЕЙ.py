from tkinter import filedialog
import tkinter.messagebox as mb
import openpyxl
import pandas as pd

msg = "Выберите выписку"
mb.showinfo("информация", msg)
filename = filedialog.askopenfilename()
df = pd.read_excel(filename, sheet_name=0, engine='openpyxl', usecols='A, B, C')
msg = "Выберите файл платников из Скарифа"
mb.showinfo("информация", msg)
filename1 = filedialog.askopenfilename()
df_base = pd.read_excel(filename1, sheet_name=0, engine='openpyxl', usecols='A, D, E')

print(df_base)
receivers_list = df_base['Получатель'].to_list()
parcels_list = df_base['Номер отправления'].to_list()

wb = openpyxl.load_workbook(filename)
ws = wb.active

dict_result = {}
i = 0
for row in df_base.iterrows():
    try:
        i += 1
        print(i)
        parcel_numb = row[1].values[0]
        receiver = row[1].values[1].replace('ё', 'е')
        pay = row[1].values[2]
        n = 0
        for cell in ws['A']:
            n += 1
            if parcel_numb in ws[f'C{n}'].value:
                print('ok')
                dict_result[parcel_numb] = receiver, ws[f'B{n}'].value
                ws[f'D{n}'].value = parcel_numb
                ws[f'E{n}'].value = cell.value
                ws[f'F{n}'].value = receiver
            elif receiver.replace('ё', 'е').lower() in cell.value.replace('ё', 'е').lower():
                print('ok2')
                dict_result[parcel_numb] = receiver, ws[f'B{n}'].value
                ws[f'D{n}'].value = parcel_numb
                ws[f'E{n}'].value = cell.value
                ws[f'F{n}'].value = receiver

    except:
        pass
wb.save(f'{filename}.xlsx')
print(dict_result)

df_dict_result = pd.DataFrame(dict_result).reset_index().transpose()
writer = pd.ExcelWriter('df_dict_result.xlsx', engine='xlsxwriter')
df_dict_result.to_excel(writer, sheet_name='Sheet1')
writer.save()
