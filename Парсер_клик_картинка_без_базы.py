import webbrowser
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from tkinter import *
import time
from threading import Thread
import keyboard
import pyautogui as pag
import pyperclip
import tkinter.messagebox as mb
import openpyxl
#import mss.tools
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment

"""sr = speech_recognition.Recognizer()
sr.pause_threshold = 1"""

#fileName = filedialog.askopenfilename()
fileName = filedialog.askopenfilename()



try:
    df_list = pd.read_excel(fileName, sheet_name='На_перевод', engine='openpyxl', header=None, usecols='L, N', skiprows=1)
except:
    wb0 = openpyxl.load_workbook(fileName)
    ws0 = wb0.worksheets[0]
    ws0.title = "На_перевод"
    wb0.save(f'{fileName}')
    df_list = pd.read_excel(fileName, sheet_name='На_перевод', engine='openpyxl', header=None, usecols='L, N',
                            skiprows=1)
#df_list = df_list.drop_duplicates()

china_names = df_list[df_list.columns[0]].tolist()
urls = df_list[df_list.columns[1]].tolist()

print(china_names)
print(urls)
print(len(df_list))
label = 1
label_3 = 1
def shift_click():
    global running
    if running == True:
        running = False
    else:
        running = True


def window():
    global label
    top = tk.Tk()
    Button = tk.Button(top, text="Start", width=8, command=shift_click, height=2)
    Button.configure(font=('Times', 20))
    lable_main = Label(top, text='Парсер из заголовков ссылки (Яндекс.браузер)', fg='black')
    lable_main.pack()
    Button.pack()
    label_0 = Label(top, text='№ Ссылки:', fg='black')
    label_0.pack()
    label = Label(top, text='', fg='black', font=('Times', 30))
    label.pack()
    label_2 = Label(top, text='время задержки:', fg='black')
    label_2.pack()
    entry = Entry(top)
    entry.pack()
    def get_profit():
        global profit
        profit = int(entry.get())
    button_take_time = tk.Button(top, text="Установить", command=get_profit)
    button_take_time.configure(font=('Times', 10))
    button_take_time.pack()
    top.mainloop()

keyboard.add_hotkey('shift', shift_click)
running = False

t = Thread(target=window)
t.start()

list_delete = [' - купить по выгодной цене в интернет-магазине OZON',
               ' — купить в интернет-магазине OZON с быстрой доставкой',
               ' по низкой цене: отзывы, фото, характеристики в интернет-магазине Ozon',
               ' - купить на OZON',
               '- купить по доступным ценам в интернет-магазине OZON',
               ' купить по выгодным ценам в интернет-магазине OZON',
               ' купить по низкой цене с доставкой в интернет-магазине OZON',
               'купить по низкой цене в интернет-магазине OZON',
               'купить по выгодной цене в интернет-магазине OZON',
               ' купить по доступной цене с доставкой в интернет-магазине OZON',
               ' купить по низкой цене в интернет-магазине OZON',
               ' — купить в интернет-магазине OZON',
               '. Итальянская мода (журнал)',
               'Купить ',
               ' - в интернет-магазине OZON с доставкой по России',
               ' - по низким ценам в интернет-магазине OZON',
               ' - с доставкой по выгодным ценам в интернет-магазине OZON',
               ' по выгодной цене в интернет-магазине OZON',
               ' по низкой цене с доставкой в интернет-магазине OZON',
               ' - по доступной цене c доставкой в интернет-магазине OZON',
                'по низкой цене в интернет-магазине OZON с доставкой',
               ' в интернет-магазине OZON по выгодной цене',
               ' - в интернет-магазине OZON по выгодной цене',
               'купить по низким ценам в интернет-магазине OZON',
               ' - купить в интернет-магазине OZON с доставкой по России',
               'купить по низким ценам с доставкой в интернет-магазине OZON']

i = 0
list_itemDescriptions = []
wb = openpyxl.load_workbook(fileName)
ws = wb.create_sheet("Sheet_A")
ws.title = "с_картинкой"
ws = wb['с_картинкой']
for url, name in zip(urls, china_names):
    if running == False:
        while running == False:
            # ожидаем повторного нажатия кнопочки
            time.sleep(2)
    if running == True:
        # выполняем основную полезную работу программы
        i += 1
        print(i)
        label.config(text=i, fg='black', font=("Times", 20))
        try:
            webbrowser.open(url)
            time.sleep(profit)
            pag.click(x=580, y=55, button='right')
            time.sleep(1)
            pag.click(x=605, y=90)
            header_text = pyperclip.paste()
            if 'Кулинарный инструмент' in header_text:
                pag.click(x=318, y=333)
                time.sleep(profit)
                pag.click(x=580, y=55, button='right')
                time.sleep(1)
                pag.click(x=605, y=90)
                header_text = pyperclip.paste()
            for elem in list_delete:
                if elem in header_text:
                    header_text = header_text.replace(elem, '')
            pag.screenshot(f'{i}.png', region=(200, 200, 900, 800))
            """with mss.mss() as sct:
                # The screen part to capture
                monitor = {"0top": 400, "left": 400, "width": 400, "height": 400}
                output = f"{i}.png"
                # Grab the data
                sct_img = sct.grab(monitor)
                # Save to the picture file
                mss.tools.to_png(sct_img.rgb, sct_img.size, output=output)
                print(output)"""

            img = f'{i}.png'

            im = Image.open(img)
            im = im.convert("RGB")
            im = im.resize((430, 190), Image.ANTIALIAS)
            im.save(f"{i}.jpg", optimize=True, quality=60)

            time.sleep(1)
            list_itemDescriptions.append(header_text)
            img = openpyxl.drawing.image.Image(f'{i}.jpg')
            img.height = 190 # insert image height in pixels as float or int (e.g. 305.5)
            img.width = 430 # insert image width in pixels as float or int (e.g. 405.8)
            row_num = i
            cell_addr = f"A{row_num}"
            img.anchor = cell_addr
            ws.add_image(img)
            ws[f"C{row_num}"] = name
            ws[f"C{row_num}"].alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
            ws[f"D{row_num}"].hyperlink = url
            ws[f"D{row_num}"].value = url
            ws[f"D{row_num}"].style = "Hyperlink"
            ws[f"D{row_num}"].alignment = Alignment(horizontal='center')
            ws[f"B{row_num}"] = header_text
            ws[f"B{row_num}"].alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
            ws.row_dimensions[row_num].height = int(180 * .8)
            ws.column_dimensions["A"].width = int(300 * .2)
            pag.click(x=189, y=15)
            pag.hotkey('ctrl', 'w')
        except:
            list_itemDescriptions.append('Error')

wb.save(f'{fileName}-с картинкой.xlsx')
msg = "Смотри файл 'output' в папке с проектом"
mb.showinfo("важная информация", msg)
print('Готово')
input()
