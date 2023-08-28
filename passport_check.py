import pandas as pd
import sqlite3
from tkinter import filedialog
import tkinter.messagebox as mb
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import openpyxl

fileName = filedialog.askopenfilename()

df = pd.read_excel(fileName, usecols='A, O, P', engine='openpyxl', header=None, skiprows=1)
print(df)
df['трек'] = df[0]
df['Серия паспорта'] = df[14]
df['Номер паспорта'] = df[15]

df['Серия паспорта'] = pd.to_numeric(df['Серия паспорта'], errors='coerce').fillna(0).astype('int64')
df['Номер паспорта'] = pd.to_numeric(df['Номер паспорта'], errors='coerce').fillna(0).astype('int64')
print(type(df['Серия паспорта'].values[1]))
print(df['Серия паспорта'])

df['Серия паспорта'] = df['Серия паспорта'].astype('str')
df['Номер паспорта'] = df['Номер паспорта'].astype('str')
df_to_baza = df.drop_duplicates()
df_to_baza = df_to_baza[['Серия паспорта', 'Номер паспорта']]
df_to_baza = df_to_baza.rename(columns={'Серия паспорта': 'series', 'Номер паспорта': 'number'})


con = sqlite3.connect(r'C:\Users\User\PycharmProjects\Переводчик\PASS_BAZA.db')
cur = con.cursor()
with con:
    baza = con.execute("select count(*) from sqlite_master where type='table' and name='pass'")
    for row in baza:
        # если таких таблиц нет
        if row[0] == 0:
            # создаём таблицу
            with con:
                con.execute("""
                            CREATE TABLE pass (
                            ID INTEGER PRIMARY KEY AUTOINCREMENT,
                            series VARCHAR(4),
                            number VARCHAR(6)
                            );
                        """)
    df_to_baza.to_sql('pass', con, if_exists='replace', index=False)

    names = [description[0] for description in con.execute("Select * from pass").description]
print(names)
con.commit()


# drop table
#con = sqlite3.connect(r'C:\Users\Илья\PycharmProjects\\Переводчик\save_pandas.db')
#with con:
#    con.executescript('drop table if exists pass')

with con:
    # Query for INNER JOIN
    sql = '''SELECT PASS_BAZA.PASSP_SERIES, PASS_BAZA.PASSP_NUMBER 
    FROM PASS_BAZA 
    INNER JOIN pass
    ON PASS_BAZA.PASSP_SERIES = pass.series
    AND PASS_BAZA.PASSP_NUMBER = pass.number'''
    df_finish = pd.read_sql(sql, con)

con.commit()
con.close()

df['ID'] = df['Серия паспорта'] + df['Номер паспорта']
df_finish['ID'] = df_finish['PASSP_SERIES'] + df_finish['PASSP_NUMBER']
df_merged = pd.merge(df, df_finish, how='left', left_on='ID' , right_on='ID')
df_merged_to_select = df_merged[df_merged['PASSP_SERIES'].notnull()]
df_merged_to_select = df_merged_to_select.drop_duplicates()
df_merged_to_select = df_merged_to_select.dropna(how='any', axis=0)
df_merged_to_select = df_merged_to_select[df_merged_to_select['PASSP_SERIES'] != '0']

print(df_merged_to_select)
if not df_merged_to_select.empty:
    redFill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')
    wb = openpyxl.load_workbook(fileName)
    ws = wb.active
    parcel_list = df_merged_to_select[0].to_list()
    i = 0
    for cell in ws['A']:
        i += 1
        if cell.value in parcel_list:
            ws[f'O{i}'].fill = redFill
            ws[f'P{i}'].fill = redFill
            print('ok')
    wb.save(fileName)

    writer = pd.ExcelWriter(f'{fileName} - список недействительных.xlsx', engine='openpyxl')
    df_merged_to_select.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
    writer.save()

wb = openpyxl.load_workbook(fileName)
ws = wb.active
#df_to_check = pd.read_excel(fileName, usecols='L', engine='openpyxl', header=None, skiprows=1)
df_exclude = pd.read_excel('exclude.xlsx', engine='openpyxl', converters={'Исключить': str, 'Убрать': str,
                                                                          'Заменить': str, 'ЗаменитьНа': str})

exclude_list = df_exclude['Исключить'].to_list()
change_list = df_exclude['Убрать'].to_list()
change_list = [item for item in change_list if not(pd.isnull(item)) == True]
change_list_2 = df_exclude['Заменить'].to_list()
change_list_2 = [item for item in change_list_2 if not(pd.isnull(item)) == True]
change_list_3 = df_exclude['ЗаменитьНа'].to_list()
change_list_3 = [item for item in change_list_3 if not(pd.isnull(item)) == True]

redFill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')
i = 0
for cell in ws['L']:
    i += 1
    numb = 0
    for item in exclude_list:
        if item in cell.value.lower():
            cell.fill = redFill
            print(item)
    for item_ch in change_list:
        if item_ch in cell.value.lower():
            cell.value = cell.value.lower().replace(item_ch, '')
            print(item_ch)
    for item_ch_2 in change_list_2:
        if item_ch_2 in cell.value.lower():
            cell.value = cell.value.lower().replace(item_ch_2, change_list_3[numb])
            print(item_ch_2)
        numb += 1
wb.save(fileName)

msg = "ГОТОВО!"
mb.showinfo("ИНФО", msg)